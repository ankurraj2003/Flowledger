"""
=============================================================
Converts validated Invoice/PO JSON into a professionally
formatted Excel file suitable for Tally / Zoho ERP import.
Supports batch export — one row per invoice in a ledger.
=============================================================
"""

import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger("flowledger.exporter")

# ─── Style Constants ─────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2C3E50")
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=10, color="7F8C8D")
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", bold=True, size=11, color="E74C3C")
BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
ACCENT_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def create_batch_excel_export(invoices: list[dict]) -> io.BytesIO:
    """
    Generate a consolidated Excel workbook from multiple invoices.
    Each invoice occupies one row in a ledger-style sheet.

    Args:
        invoices: List of validated invoice dictionaries,
                  each with a 'file_name' key for source tracking.

    Returns:
        BytesIO buffer containing the Excel file.
    """
    logger.info(f"Generating batch Excel export for {len(invoices)} invoice(s)...")
    wb = Workbook()

    # ─── Sheet 1: Invoice Ledger (one row per invoice) ───────
    ws_ledger = wb.active
    ws_ledger.title = "Invoice Ledger"
    ws_ledger.sheet_properties.tabColor = "2C3E50"

    # Title
    ws_ledger.merge_cells("A1:N1")
    cell_title = ws_ledger["A1"]
    cell_title.value = "📄 INVOICE LEDGER — Flowledger Batch Export"
    cell_title.font = TITLE_FONT

    ws_ledger.merge_cells("A2:N2")
    cell_sub = ws_ledger["A2"]
    cell_sub.value = (
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        f"{len(invoices)} invoice(s) processed | Flowledger V1.0"
    )
    cell_sub.font = SUBTITLE_FONT

    # Column headers at row 4
    columns = [
        "#",                    # A
        "Source File",          # B
        "Invoice No.",          # C
        "Vendor Name",          # D
        "Vendor Address",       # E
        "Invoice Date",         # F
        "Due Date",             # G
        "Bill To",              # H
        "Ship To",              # I
        "Payment Terms",        # J
        "Items (Count)",        # K
        "Items Description",    # L
        "Sub Total ($)",        # M
        "Tax Rate (%)",         # N
        "Tax Amount ($)",       # O
        "Grand Total ($)",      # P
    ]
    col_widths = [5, 22, 16, 24, 30, 14, 14, 28, 28, 16, 12, 50, 14, 12, 14, 16]

    for col_idx, (header, width) in enumerate(zip(columns, col_widths), start=1):
        cell = ws_ledger.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws_ledger.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows — one per invoice
    for i, inv in enumerate(invoices, start=1):
        row = i + 4  # Starts at row 5
        items = inv.get("items", [])

        # Build a summary of all line items for one cell
        item_descriptions = "; ".join(
            f"{it.get('desc', '')} (x{it.get('qty', 0)})"
            for it in items
        )

        values = [
            i,                                          # #
            inv.get("file_name", "N/A"),                # Source File
            inv.get("invoice_no", "N/A"),               # Invoice No.
            inv.get("vendor_name", "N/A"),               # Vendor Name
            inv.get("vendor_address", "N/A"),            # Vendor Address
            inv.get("date", "N/A"),                      # Invoice Date
            inv.get("due_date", "N/A"),                  # Due Date
            inv.get("bill_to", "N/A"),                   # Bill To
            inv.get("ship_to", "N/A"),                   # Ship To
            inv.get("terms", "N/A"),                     # Payment Terms
            len(items),                                  # Items Count
            item_descriptions,                           # Items Description
            inv.get("sub_total", 0),                     # Sub Total
            inv.get("tax_rate", 0),                      # Tax Rate
            inv.get("tax_amount", 0),                    # Tax Amount
            inv.get("grand_total", 0),                   # Grand Total
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws_ledger.cell(row=row, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = BORDER

            if col_idx == 1:  # Row number
                cell.alignment = CENTER
            elif col_idx in (13, 14, 15, 16):  # Numeric cols
                cell.alignment = RIGHT
                if col_idx != 14:  # Not tax rate
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.00"%"'
            elif col_idx == 12:  # Items description
                cell.alignment = LEFT
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Alternating row colors
            if i % 2 == 0:
                cell.fill = ACCENT_FILL

    # Grand total row
    if invoices:
        total_row = len(invoices) + 5
        ws_ledger.merge_cells(f"A{total_row}:L{total_row}")
        cell_label = ws_ledger.cell(row=total_row, column=1, value=f"TOTAL ({len(invoices)} invoices)")
        cell_label.font = TOTAL_FONT
        cell_label.alignment = Alignment(horizontal="right", vertical="center")
        cell_label.border = BORDER

        # Sum of Sub Total
        sub_sum = sum(inv.get("sub_total", 0) for inv in invoices)
        cell = ws_ledger.cell(row=total_row, column=13, value=sub_sum)
        cell.font = TOTAL_FONT
        cell.alignment = RIGHT
        cell.number_format = '#,##0.00'
        cell.border = BORDER

        # Sum of Tax
        tax_sum = sum(inv.get("tax_amount", 0) for inv in invoices)
        cell = ws_ledger.cell(row=total_row, column=15, value=tax_sum)
        cell.font = TOTAL_FONT
        cell.alignment = RIGHT
        cell.number_format = '#,##0.00'
        cell.border = BORDER

        # Sum of Grand Total
        grand_sum = sum(inv.get("grand_total", 0) for inv in invoices)
        cell = ws_ledger.cell(row=total_row, column=16, value=grand_sum)
        cell.font = TOTAL_FONT
        cell.alignment = RIGHT
        cell.number_format = '#,##0.00'
        cell.border = BORDER

    # ─── Sheet 2: Detailed Line Items (all invoices) ─────────
    ws_items = wb.create_sheet(title="All Line Items")
    ws_items.sheet_properties.tabColor = "27AE60"

    ws_items.merge_cells("A1:I1")
    ws_items["A1"].value = "📋 ALL LINE ITEMS — Detailed Breakdown"
    ws_items["A1"].font = TITLE_FONT

    item_columns = [
        "#", "Source File", "Invoice No.", "Item Description",
        "Details", "Internal SKU", "Qty", "Unit", "Rate ($)", "Amount ($)"
    ]
    item_widths = [5, 20, 16, 30, 28, 18, 10, 12, 14, 14]

    for col_idx, (header, width) in enumerate(zip(item_columns, item_widths), start=1):
        cell = ws_items.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws_items.column_dimensions[get_column_letter(col_idx)].width = width

    # Flatten all line items across all invoices
    row_num = 0
    for inv in invoices:
        for item in inv.get("items", []):
            row_num += 1
            row = row_num + 3
            sku = item.get("internal_sku", "MANUAL REVIEW")

            values = [
                row_num,
                inv.get("file_name", ""),
                inv.get("invoice_no", ""),
                item.get("desc", ""),
                item.get("details", ""),
                sku,
                item.get("qty", 0),
                item.get("unit", "Piece"),
                item.get("price", 0),
                item.get("total", 0),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws_items.cell(row=row, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = BORDER
                if col_idx in (9, 10):
                    cell.alignment = RIGHT
                    cell.number_format = '#,##0.00'
                elif col_idx in (1, 7):
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT

                if row_num % 2 == 0:
                    cell.fill = ACCENT_FILL

    # ─── Write to buffer ─────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(f"Batch export generated — {len(invoices)} invoices, {row_num} total line items.")
    return buffer
